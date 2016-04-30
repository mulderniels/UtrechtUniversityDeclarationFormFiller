#made by n.mulder1@uu.nl 2016

#fills Utrecht University declaration form C1 (template.xlsx) for each row in data.xlsx

#init
import os.path
from openpyxl import load_workbook

#do stuff
print('hallo')

data = load_workbook(filename = 'data.xlsx')
data = data.active #select the active sheet

personsN = data.max_row-1 #no of scenarios

templateFile = load_workbook(filename = 'template.xlsx')
template = templateFile.active #select the active sheet

#loop trough persons
for iPerson in range(2, personsN+1+1):
	templateFile = load_workbook(filename = 'template.xlsx')
	result = templateFile.active #select the active sheet
	
	achternaam = str(data.cell(row = iPerson, column = 3).value);
	filename = achternaam + ' ' + str(iPerson);
	
	result['A11'] = str(data.cell(row = iPerson, column = 2).value)[0] #voorletter
	result['J11'] = achternaam #achternaam
	result['A13'] = str(data.cell(row = iPerson, column = 6).value) #straat
	result['J13'] = str(data.cell(row = iPerson, column = 7).value) #postcode
	result['A15'] = str(data.cell(row = iPerson, column = 4).value) #iban
	result['A17'] = str(data.cell(row = iPerson, column = 5).value) #swift
	if not os.path.exists('output/' + filename):
	    os.makedirs('output/' + filename)
	templateFile.save('output/' + filename + '/declaratie ' + filename + '.xlsx')

print('doei')